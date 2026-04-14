#!/usr/bin/env python3
"""
ODE G2M — McKinsey-style Excel workbook generator
One sheet per dashboard + Executive Summary
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE

# ─── PALETTE ────────────────────────────────────────────────────────────────
C_NAVY      = "0C1F3F"   # section headers
C_INDIGO    = "635BFF"   # accent / highlight
C_INDIGO_LT = "EEEDFF"   # light indigo tint
C_SLATE     = "2D3055"   # sub-headers
C_MIDGRAY   = "6B6F8E"   # helper text
C_RULE      = "E2E3EA"   # thin divider lines
C_STRIPE    = "F8F8FC"   # alternating row fill
C_WHITE     = "FFFFFF"
C_BLACK     = "0C0E1A"

# ─── STYLE HELPERS ──────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border(style="thin", color="D0D3E0"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border(color="D0D3E0"):
    n = Side(style=None)
    b = Side(style="thin", color=color)
    return Border(left=n, right=n, top=n, bottom=b)

def font(name="Calibri", size=10, bold=False, italic=False, color=C_BLACK):
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def apply(ws, row, col, value, fnt=None, fll=None, brd=None, aln=None, num_fmt=None):
    cell = ws.cell(row=row, column=col)
    try:
        cell.value = value
    except AttributeError:
        return cell  # MergedCell slave — skip
    if fnt: cell.font = fnt
    if fll: cell.fill = fll
    if brd: cell.border = brd
    if aln: cell.alignment = aln
    if num_fmt: cell.number_format = num_fmt
    return cell

def merge_write(ws, r1, c1, r2, c2, value, fnt=None, fll=None, brd=None, aln=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    if fnt: cell.font = fnt
    if fll: cell.fill = fll
    if aln: cell.alignment = aln
    return cell

def page_title(ws, row, title, subtitle=None, ncols=10):
    """Big page title block at top of each sheet"""
    merge_write(ws, row, 1, row, ncols, title,
        fnt=font("Calibri", 16, bold=True, color=C_NAVY),
        fll=fill(C_WHITE),
        aln=align("left", "center"))
    ws.row_dimensions[row].height = 30
    row += 1
    if subtitle:
        merge_write(ws, row, 1, row, ncols, subtitle,
            fnt=font("Calibri", 10, italic=True, color=C_MIDGRAY),
            aln=align("left", "center"))
        ws.row_dimensions[row].height = 16
        row += 1
    # thin rule
    for c in range(1, ncols+1):
        ws.cell(row=row, column=c).border = bottom_border(C_INDIGO)
    ws.row_dimensions[row].height = 4
    return row + 1

def section_header(ws, row, label, ncols=10):
    """Dark navy section divider"""
    merge_write(ws, row, 1, row, ncols, f"  {label}",
        fnt=font("Calibri", 9, bold=True, color=C_WHITE),
        fll=fill(C_NAVY),
        aln=align("left", "center"))
    ws.row_dimensions[row].height = 18
    return row + 1

def sub_header(ws, row, label, ncols=10):
    """Slate sub-section label"""
    merge_write(ws, row, 1, row, ncols, f"  {label}",
        fnt=font("Calibri", 9, bold=True, color=C_WHITE),
        fll=fill(C_SLATE),
        aln=align("left", "center"))
    ws.row_dimensions[row].height = 16
    return row + 1

def table_header(ws, row, cols, widths=None):
    """Column header row for a data table"""
    for i, col in enumerate(cols, 1):
        apply(ws, row, i, col,
            fnt=font("Calibri", 9, bold=True, color=C_WHITE),
            fll=fill(C_SLATE),
            brd=border("thin", C_RULE),
            aln=align("center", "center", wrap=True))
    ws.row_dimensions[row].height = 30
    return row + 1

def table_row(ws, row, values, stripe=False, bold_first=False):
    """Data row, optionally striped"""
    bg = C_STRIPE if stripe else C_WHITE
    for i, val in enumerate(values, 1):
        b = bold_first and i == 1
        apply(ws, row, i, val,
            fnt=font("Calibri", 9, bold=b, color=C_BLACK),
            fll=fill(bg),
            brd=border("thin", C_RULE),
            aln=align("left", "center", wrap=True))
    ws.row_dimensions[row].height = 24
    return row + 1

def kpi_row(ws, row, kpis, ncols=10):
    """Stats bar — wide KPI cells across the sheet"""
    n = len(kpis)
    span = ncols // n
    col = 1
    for num, lbl in kpis:
        c2 = col + span - 1
        if col == (n-1)*span + 1:  # last cell gets remainder
            c2 = ncols
        ws.merge_cells(start_row=row, start_column=col, end_row=row+1, end_column=c2)
        cell = ws.cell(row=row, column=col, value=num)
        cell.font = Font(name="Calibri", size=18, bold=True, color=C_INDIGO)
        cell.fill = fill(C_INDIGO_LT)
        cell.alignment = Alignment(horizontal="center", vertical="bottom")
        ws.merge_cells(start_row=row+2, start_column=col, end_row=row+2, end_column=c2)
        cell2 = ws.cell(row=row+2, column=col, value=lbl.upper())
        cell2.font = Font(name="Calibri", size=7.5, bold=False, color=C_MIDGRAY)
        cell2.fill = fill(C_INDIGO_LT)
        cell2.alignment = Alignment(horizontal="center", vertical="top")
        # right border as divider
        brd_cell = ws.cell(row=row, column=c2)
        brd_cell.border = Border(right=Side(style="thin", color=C_RULE))
        col = c2 + 1
    ws.row_dimensions[row].height = 26
    ws.row_dimensions[row+1].height = 6
    ws.row_dimensions[row+2].height = 18
    return row + 4

def spacer(ws, row, height=8):
    ws.row_dimensions[row].height = height
    return row + 1

def insight_row(ws, row, label, text, ncols=10):
    """Callout/insight box"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    apply(ws, row, 1, label.upper(),
        fnt=font("Calibri", 8, bold=True, color=C_INDIGO),
        fll=fill(C_INDIGO_LT),
        aln=align("left", "center"))
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=ncols)
    apply(ws, row, 3, text,
        fnt=font("Calibri", 9, italic=True, color=C_SLATE),
        fll=fill(C_INDIGO_LT),
        aln=align("left", "center", wrap=True))
    ws.row_dimensions[row].height = 28
    return row + 1

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, row=4, col=1):
    ws.freeze_panes = ws.cell(row=row, column=col)

# ─── SHEET 1: EXECUTIVE SUMMARY ─────────────────────────────────────────────
def build_summary(wb):
    ws = wb.create_sheet("Executive Summary", 0)
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [3, 22, 18, 18, 18, 14, 14, 14, 14, 14])

    r = 2
    r = page_title(ws, r, "ODE — Go-To-Market Strategy Hub",
        "Executive Summary  ·  April 2026  ·  Internal Confidential")

    r = spacer(ws, r, 6)
    r = section_header(ws, r, "01  COMPANY OVERVIEW")
    r = spacer(ws, r, 4)

    overview = [
        ("Company", "Orthogonal Design Engineering (ODE)"),
        ("Product", "ode. — AI-native full-stack engineering platform"),
        ("Positioning", '"The Dassault of the AI era" — AI-native CAE covering 12 engineering disciplines'),
        ("Target Market", "Global CAE/simulation market: $29B TAM (2025, expanded)"),
        ("Stage", "Pre-revenue · Pre-A round · Building to $1M ARR (12 months)"),
        ("HQ", "EU-based · Global GTM from Day 1"),
    ]
    for label, val in overview:
        apply(ws, r, 2, label, fnt=font("Calibri", 9, bold=True, color=C_SLATE),
            aln=align("left", "center"))
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
        apply(ws, r, 3, val, fnt=font("Calibri", 9, color=C_BLACK),
            aln=align("left", "center", wrap=True))
        ws.row_dimensions[r].height = 20
        r += 1

    r = spacer(ws, r, 8)
    r = section_header(ws, r, "02  12-MONTH FINANCIAL TARGETS")
    r = spacer(ws, r, 4)

    r = table_header(ws, r, ["", "KPI", "Q1", "Q2", "Q3", "Q4 (Year-End)", "", "", "", ""])
    kpis = [
        ("", "ARR Target",           "$50K",   "$216K",  "$588K",  "$1,000,000"),
        ("", "MRR",                   "$3.3K",  "$18K",   "$49K",   "$83K+"),
        ("", "Paid Users",            "15",     "65",     "150",    "280"),
        ("", "Free Sign-ups",         "500",    "1,500",  "2,800",  "4,500"),
        ("", "Enterprise Accounts",   "0",      "3",      "6",      "12–15"),
        ("", "Token Revenue Share",   "—",      "20%",    "35%",    "40%+"),
        ("", "Net Revenue Retention", "—",      "105%",   "115%",   "120%+"),
        ("", "Channel Partners",      "0",      "1",      "2",      "3–5"),
        ("", "Organic Traffic / mo",  "2,000",  "5,000",  "7,500",  "10,000"),
        ("", "Blended CAC",           "—",      "—",      "—",      "<$800"),
    ]
    for i, row_data in enumerate(kpis):
        ws_row = table_row(ws, r, list(row_data[:6]) + ["","","",""], stripe=(i%2==1))
        r = ws_row

    r = spacer(ws, r, 8)
    r = section_header(ws, r, "03  DASHBOARD INDEX")
    r = spacer(ws, r, 4)

    r = table_header(ws, r, ["", "Dashboard", "Description", "Key Sections", "Linked Sheet", "", "", "", "", ""])
    dashboards = [
        ("", "GTM Strategy",       "Market opportunity, ICP, revenue model, 12-month roadmap, channels, KPIs",
         "Market · ICP · Revenue · Roadmap · Channels · KPIs",          "GTM Strategy"),
        ("", "Sales Playbook",     "Pipeline funnel, BANT+ qualification, demo playbook, objections, ROI",
         "Pipeline · Qualification · Demo · Objections · ROI",           "Sales Playbook"),
        ("", "Brand Strategy",     "Positioning, competitive landscape, brand architecture, messaging matrix",
         "Positioning · Competitive · Architecture · Messaging",         "Brand Strategy"),
        ("", "Marketing Strategy", "Content engine, channel mix, email sequences, SEO, webinar calendar",
         "Content · Channels · Funnel · Email · SEO · Webinars",         "Marketing Strategy"),
    ]
    for i, (a, name, desc, sections, sheet) in enumerate(dashboards):
        bg = C_STRIPE if i%2 else C_WHITE
        for c, val in enumerate(["", name, desc, sections, sheet, "","","","",""], 1):
            bold = c == 2
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=bold, color=C_INDIGO if bold else C_BLACK),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left", "center", wrap=True))
        ws.row_dimensions[r].height = 32
        r += 1

    r = spacer(ws, r, 8)
    r = section_header(ws, r, "04  REVENUE MODEL SUMMARY")
    r = spacer(ws, r, 4)

    r = table_header(ws, r, ["", "Tier", "Price / mo", "Token Bundle", "Target ICP", "Year-1 ARR", "", "", "", ""])
    tiers = [
        ("", "Free",       "$0",          "500 tokens / day",         "Students, explorers",                   "—"),
        ("", "Maker",      "$9–$29",       "Maker CAD + 100K tokens/mo","LEGO builders, Arduino, RPi, 3D print", "New stream"),
        ("", "Starter",    "$49",          "500K tokens / mo",         "Indie engineers",                       "$150K (15%)"),
        ("", "Pro",        "$99",          "5M tokens / mo",           "Consultancies, professionals",          "$250K (25%)"),
        ("", "Team",       "$199/seat",    "20M tokens / seat",        "Engineering teams",                     "—"),
        ("", "Enterprise", "Custom",       "Dedicated pool",           "OEMs, large orgs",                      "$450K (45%)"),
    ]
    for i, row_data in enumerate(tiers):
        r = table_row(ws, r, list(row_data[:6]) + ["","","",""], stripe=(i%2==1))

    freeze(ws, 4)
    return ws

# ─── SHEET 2: GTM STRATEGY ───────────────────────────────────────────────────
def build_gtm(wb):
    ws = wb.create_sheet("GTM Strategy")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [3, 22, 16, 16, 16, 14, 14, 12, 12, 10])

    r = 2
    r = page_title(ws, r, "GTM Strategy — Market Opportunity, ICP & Revenue Model",
        "ODE · April 2026 · Internal Confidential")
    r = spacer(ws, r, 4)

    # KPIs
    kpis = [("$1M","Year 1 ARR Target"),("$29B","Global TAM"),
            ("280","Paid Users (M12)"),("40/60","PLG / SLG Split"),("$3,500","Blended ARPU / yr")]
    r = kpi_row(ws, r, kpis)
    r = spacer(ws, r, 6)

    # 01 MARKET
    r = section_header(ws, r, "01  MARKET OPPORTUNITY")
    r = spacer(ws, r, 3)

    r = insight_row(ws, r, "Why Now",
        "Four structural forces converge: LLMs reaching engineering competence · Engineering AI budgets mandated 2025–26 · "
        "Dassault's 30-year binary architecture cannot adapt · Cloud/browser expectations normalized among engineers.")
    r = spacer(ws, r, 4)

    r = sub_header(ws, r, "TAM / SAM / SOM")
    r = table_header(ws, r, ["", "Market", "Size", "Definition", "", "", "", "", "", ""])
    tam = [
        ("", "TAM (Expanded)",  "$29B",   "Engineering simulation ($23B) + Maker/IoT tools ($3.5B) + 3D print services ($2.8B)"),
        ("", "SAM",             "$1.3B",  "AI-addressable engineering tools segment"),
        ("", "Beachhead",       "$800M",  "AI-native tools subsegment, 45%+ CAGR"),
        ("", "SOM Year 1",      "$1.2M",  "Realistic capture across all 5 segments"),
    ]
    for i, (a, mkt, size, defn) in enumerate(tam):
        bg = C_STRIPE if i%2 else C_WHITE
        for c, val in enumerate(["", mkt, size, defn, "","","","","",""], 1):
            bold = c in (2, 3)
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=bold, color=C_SLATE if bold else C_BLACK),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left", "center", wrap=True))
        ws.row_dimensions[r].height = 22
        r += 1
    r = spacer(ws, r, 4)

    r = sub_header(ws, r, "SOM Year-1 by Segment")
    r = table_header(ws, r, ["", "Segment", "ARR Target", "% of SOM", "", "", "", "", "", ""])
    som = [
        ("", "Enterprise OEMs",     "$450K", "37.5%"),
        ("", "Consultancies",        "$250K", "20.8%"),
        ("", "Universities",         "$150K", "12.5%"),
        ("", "Indie Engineers",      "$150K", "12.5%"),
        ("", "Channel Partners",     "$200K", "16.6%"),
        ("", "Total",                "$1.2M", "100%"),
    ]
    for i, (a, seg, arr, pct) in enumerate(som):
        bold = seg == "Total"
        bg = C_NAVY if bold else (C_STRIPE if i%2 else C_WHITE)
        c_txt = C_WHITE if bold else C_BLACK
        for c, val in enumerate(["", seg, arr, pct, "","","","","",""], 1):
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=bold, color=c_txt),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left" if c==2 else "center", "center"))
        ws.row_dimensions[r].height = 20
        r += 1
    r = spacer(ws, r, 8)

    # 02 ICP MATRIX
    r = section_header(ws, r, "02  ICP MATRIX")
    r = spacer(ws, r, 3)

    r = table_header(ws, r, ["", "Segment", "Size", "Target ACV", "Sales Cycle", "Motion", "ARR Target", "Priority", "", ""])
    icp = [
        ("", "Enterprise OEMs",          "500–50K+", "$50K–$150K",   "6–12 months",  "SLG Direct",     "$450K (45%)", "P1"),
        ("", "Engineering Consultancies", "10–500",   "$10K–$25K",    "4–8 weeks",    "SLG Light",      "$250K (25%)", "P2"),
        ("", "Universities",             "1K–50K",   "$5K–$15K",     "3–8 months",   "SLG Academic",   "$150K (15%)", "P3"),
        ("", "Indie Engineers",          "1–50",     "$1.2K–$3K",    "Same-day–2wks","PLG Self-serve",  "$150K (15%)", "P1"),
        ("", "Channel Partners",         "5–200",    "$200K+ pipeline","30–60 days", "Partner SLG",    "$200K+",      "P0"),
        ("", "Maker & Enthusiast",       "1 (consumer)","$108–$348/yr","Same-day",  "PLG Self-serve",  "New stream",  "P2"),
    ]
    for i, row_data in enumerate(icp):
        bg = C_STRIPE if i%2 else C_WHITE
        for c, val in enumerate(list(row_data[:8]) + ["",""], 1):
            bold = c == 2
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=bold, color=C_INDIGO if c==8 else C_BLACK),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left", "center", wrap=True))
        ws.row_dimensions[r].height = 24
        r += 1
    r = spacer(ws, r, 8)

    # 03 REVENUE MODEL
    r = section_header(ws, r, "03  REVENUE MODEL")
    r = spacer(ws, r, 3)

    r = insight_row(ws, r, "Model Logic",
        "Subscriptions are the access vehicle. Tokens are the revenue engine. Every AI action consumes tokens. "
        "ARPU compounds as usage grows. Maker/consumer streams (print-on-demand, module marketplace, hardware bundles) "
        "add high-volume transactional revenue on top of SaaS base.")
    r = spacer(ws, r, 3)

    r = sub_header(ws, r, "Subscription Pricing Tiers")
    r = table_header(ws, r, ["", "Tier", "Price / mo", "Token Allocation", "Target ICP", "ARR (Yr 1)", "", "", "", ""])
    pricing = [
        ("", "Free",       "$0",         "500 tokens / day",          "Students, explorers",                    "—"),
        ("", "Maker",      "$9–$29",      "Maker CAD + 100K tokens/mo","LEGO builders, Arduino, RPi, 3D print", "New stream"),
        ("", "Starter",    "$49",         "500K tokens / mo",          "Indie engineers",                        "$150K"),
        ("", "Pro",        "$99",         "5M tokens / mo",            "Consultancies",                          "$250K"),
        ("", "Team",       "$199/seat",   "20M tokens / seat",         "Engineering teams",                      "—"),
        ("", "Enterprise", "Custom",      "Dedicated pool",            "OEMs, large orgs",                       "$450K"),
    ]
    for i, row_data in enumerate(pricing):
        r = table_row(ws, r, list(row_data[:6]) + ["","","",""], stripe=(i%2==1))
    r = spacer(ws, r, 4)

    r = sub_header(ws, r, "ARPU by Segment — Year 1 at $1M ARR")
    r = table_header(ws, r, ["", "Segment", "Paid Users", "% Users", "ARPU", "Segment ARR", "% ARR", "Notes", "", ""])
    arpu = [
        ("", "Enterprise OEMs",    "42",   "15%",  "$10,714", "$450,000",    "45%", "Services + high token consumption"),
        ("", "Consultancies",       "56",   "20%",  "$4,464",  "$250,000",    "25%", "Pro/Team, project-based usage"),
        ("", "Universities",        "70",   "25%",  "$2,143",  "$150,000",    "15%", "80% academic discount applied"),
        ("", "Indie Engineers",     "112",  "40%",  "$1,339",  "$150,000",    "15%", "PLG-driven, Starter/Pro mix"),
        ("", "Total",               "280",  "100%", "$3,571",  "$1,000,000",  "100%","Blended ARPU target"),
    ]
    for i, row_data in enumerate(arpu):
        bold = row_data[1] == "Total"
        bg = C_NAVY if bold else (C_STRIPE if i%2 else C_WHITE)
        c_txt = C_WHITE if bold else C_BLACK
        for c, val in enumerate(list(row_data[:8]) + ["",""], 1):
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=bold, color=c_txt),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left" if c in (2, 8) else "center", "center", wrap=True))
        ws.row_dimensions[r].height = 22
        r += 1
    r = spacer(ws, r, 8)

    # 04 ROADMAP
    r = section_header(ws, r, "04  12-MONTH EXECUTION ROADMAP")
    r = spacer(ws, r, 3)

    r = table_header(ws, r, ["", "Phase", "Period", "Focus", "Key Milestones", "Budget", "MRR Target", "", "", ""])
    roadmap = [
        ("", "A — Foundation",      "M1–M2", "AI-Native Ops",
         "AI tool stack for all 8 roles · Website rebuild · Token pricing page · 200 free users", "$60K", "$3.3K"),
        ("", "B — Traction",        "M2–M4", "Channel Launch",
         "PLG funnel validated (40% activation) · First channel partner · Enterprise ABM 50 accounts · 65 paid users", "$75K", "$18K"),
        ("", "C — Enterprise",      "M4–M6", "Pipeline + FEM/CFD Beta",
         "3 enterprise contracts ($150K+ ACV) · FEM/CFD beta 20 users · 3 case studies · 5 universities", "$80K", "$48K"),
        ("", "D — Channel Growth",  "M6–M9", "Token Scale",
         "3 channel partners producing pipeline · FEM/CFD launch · Token revenue 35%+ MRR · NRR >115%", "$80K", "$70K"),
        ("", "E — Series A Ready",  "M9–M12","$1M ARR",
         "$83K+ MRR · 280 paid users · Token revenue 40%+ ARR · Series A data room · Year 2 roadmap", "$75K", "$83K+"),
    ]
    for i, row_data in enumerate(roadmap):
        bg = C_STRIPE if i%2 else C_WHITE
        for c, val in enumerate(list(row_data[:7]) + ["","",""], 1):
            bold = c == 2
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=bold, color=C_SLATE if bold else C_BLACK),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left", "center", wrap=True))
        ws.row_dimensions[r].height = 38
        r += 1
    r = spacer(ws, r, 8)

    # 05 CHANNELS
    r = section_header(ws, r, "05  CHANNEL STRATEGY")
    r = spacer(ws, r, 3)

    r = table_header(ws, r, ["", "Channel", "ARR %", "Target ACV", "CAC Target", "Key Motions", "", "", "", ""])
    channels = [
        ("", "PLG — Product-Led",    "40%", "—",             "<$500",
         "Free tier → limits → upgrade · 60s onboarding · 40% activation in 7 days · 5% free→paid in 30 days"),
        ("", "Direct Sales — SLG",   "40%", "$50K–$500K",    "<$3,000",
         "AI-assisted ABM (Clay, 100 ICP/wk) · Discovery → 30-day pilot → enterprise contract · 3–6 month cycle"),
        ("", "Channel Partners",      "20%", "$200K+ pipeline","Minimal",
         "3–5 EU engineering consultancies + Dassault/ANSYS resellers · Rev share 20–30% · 10–50 clients each"),
    ]
    for i, row_data in enumerate(channels):
        bg = C_STRIPE if i%2 else C_WHITE
        for c, val in enumerate(list(row_data[:6]) + ["","","",""], 1):
            bold = c == 2
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=bold, color=C_SLATE if bold else C_BLACK),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left", "center", wrap=True))
        ws.row_dimensions[r].height = 42
        r += 1
    r = spacer(ws, r, 8)

    # 06 KPI TRACKER
    r = section_header(ws, r, "06  KPI TRACKER — QUARTERLY TARGETS")
    r = spacer(ws, r, 3)

    r = table_header(ws, r, ["", "KPI", "Q1", "Q2", "Q3", "Q4 (Year-End)", "", "", "", ""])
    kpi_table = [
        ("", "MRR",                   "$3,300",  "$18,000",  "$49,000",  "$83,000+"),
        ("", "Paid Users (cumul.)",    "15",      "65",       "150",      "280"),
        ("", "Free Sign-ups (cumul.)", "500",     "1,500",    "2,800",    "4,500"),
        ("", "Enterprise Accounts",   "0",       "3",        "6",        "12–15"),
        ("", "Token Revenue Share",   "—",       "20%",      "35%",      "40%+"),
        ("", "Net Revenue Retention", "—",       "105%",     "115%",     "120%+"),
        ("", "Channel Partners",      "0",       "1",        "2",        "3–5"),
        ("", "Organic Traffic / mo",  "2,000",   "5,000",    "7,500",    "10,000"),
        ("", "Blended CAC",           "Monitor", "Monitor",  "Monitor",  "<$800"),
    ]
    for i, row_data in enumerate(kpi_table):
        r = table_row(ws, r, list(row_data[:6]) + ["","","",""], stripe=(i%2==1))

    freeze(ws, 4)
    return ws

# ─── SHEET 3: SALES PLAYBOOK ─────────────────────────────────────────────────
def build_sales(wb):
    ws = wb.create_sheet("Sales Playbook")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [3, 22, 16, 16, 16, 14, 14, 14, 10, 10])

    r = 2
    r = page_title(ws, r, "Sales Playbook — Pipeline, Qualification & ROI Framework",
        "ODE · April 2026 · Internal Confidential")
    r = spacer(ws, r, 4)

    kpis = [("$2M","Pipeline Target (Q3)"),("$50K–$500K","Enterprise ACV"),
            ("30%","Pilot → Contract"),("120%+","Net Revenue Retention"),("<$800","Blended CAC")]
    r = kpi_row(ws, r, kpis)
    r = spacer(ws, r, 6)

    # Pipeline Funnel
    r = section_header(ws, r, "01  PIPELINE OVERVIEW — ENTERPRISE FUNNEL")
    r = spacer(ws, r, 3)

    r = table_header(ws, r, ["", "Stage", "Conversion", "Volume", "Next Action", "", "", "", "", ""])
    funnel = [
        ("", "Target Accounts",  "100%",  "50 accounts",  "Research & map to ICP — use Clay for personalised outreach"),
        ("", "Discovery Call",   "60%",   "30 calls",     "BANT+ score. Score ≥4/5 → fast-track to demo within 5 days"),
        ("", "Live Demo",        "40%",   "20 demos",     "Custom demo for their domain. Push for 30-day pilot proposal"),
        ("", "Pilot Signed",     "20%",   "10 pilots",    "Week 1 setup · Weeks 2-3 team adoption · Week 4 results review"),
        ("", "Contract Closed",  "6%",    "3+ deals",     "Negotiate enterprise service package · CSM handover"),
    ]
    for i, row_data in enumerate(funnel):
        bg = C_STRIPE if i%2 else C_WHITE
        for c, val in enumerate(list(row_data[:5]) + ["","","","",""], 1):
            bold = c == 2
            pct_color = C_INDIGO if c == 3 else (C_SLATE if bold else C_BLACK)
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=bold, color=pct_color),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left", "center", wrap=True))
        ws.row_dimensions[r].height = 28
        r += 1
    r = spacer(ws, r, 8)

    # Revenue by Role
    r = section_header(ws, r, "  REVENUE RESPONSIBILITY BY ROLE")
    r = spacer(ws, r, 3)
    r = table_header(ws, r, ["", "Role", "Revenue Target", "Motion", "Start Month", "", "", "", "", ""])
    roles = [
        ("", "CEO (Founder-led)",    "$150K",          "SLG — Top 5 enterprise",        "Immediate · 60% GTM time"),
        ("", "Head of Growth",       "$400K",          "PLG — Entire funnel",            "Hire M1"),
        ("", "Account Executive",    "$300K",          "SLG — Enterprise pipeline",      "Hire M3"),
        ("", "Customer Success",     "$150K expansion","Both — NRR >120%",               "Hire M8"),
        ("", "SDR / BDR",            "Pipeline only",  "SLG — Qualified meetings",       "Hire M6"),
        ("", "Total ARR",            "$1,000,000",     "Year-end target",                "—"),
    ]
    for i, row_data in enumerate(roles):
        bold = row_data[1] == "Total ARR"
        bg = C_NAVY if bold else (C_STRIPE if i%2 else C_WHITE)
        c_txt = C_WHITE if bold else C_BLACK
        for c, val in enumerate(list(row_data[:5]) + ["","","","",""], 1):
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=bold, color=c_txt),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left", "center", wrap=True))
        ws.row_dimensions[r].height = 22
        r += 1
    r = spacer(ws, r, 8)

    # BANT+
    r = section_header(ws, r, "02  BANT+ QUALIFICATION FRAMEWORK")
    r = spacer(ws, r, 3)
    r = insight_row(ws, r, "Scoring Rule",
        "Score 5/5 = Hot (fast-track to demo)  ·  Score 4/5 = Warm (qualify further)  ·  "
        "Score 3/5 = Lukewarm (nurture 30 days)  ·  Score <3 = Cold (remove from active pipeline)")
    r = spacer(ws, r, 3)

    r = table_header(ws, r, ["", "Criterion", "What to Assess", "Strong Signal", "Weak Signal", "", "", "", "", ""])
    bant = [
        ("", "B — Budget",    "Dedicated simulation / AI budget exists",
         "\"We have a CAD/simulation budget line\" · \"AI tools budget approved\"",
         "\"Need to find budget\" · \"I'd need to request it\""),
        ("", "A — Authority", "Speaking with decision maker or champion with access",
         "VP Engineering / CTO / Managing Director in meeting",
         "\"I need to run this by my manager\" without named access"),
        ("", "N — Need",      "Active pain around simulation speed, tool fragmentation, or AI adoption",
         "Mentions tool silos, slow iteration cycles, license cost burden",
         "\"Happy with current tools\" · \"No active projects\""),
        ("", "T — Timeline",  "Decision expected within 2 quarters",
         "\"This quarter\" · \"Q2/Q3 rollout planned\"",
         "\"Maybe next year\" · \"No immediate urgency\""),
        ("", "+ Tech Fit",    "Active simulation workflow; open to Modelica/FMI",
         "Active Modelica user · FMI/FMU requirement · 5+ simulation engineers",
         "No simulation experience · Excel-based analysis only"),
    ]
    for i, row_data in enumerate(bant):
        bg = C_STRIPE if i%2 else C_WHITE
        for c, val in enumerate(list(row_data[:5]) + ["","","","",""], 1):
            bold = c == 2
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=bold, color=C_SLATE if bold else C_BLACK),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left", "center", wrap=True))
        ws.row_dimensions[r].height = 30
        r += 1
    r = spacer(ws, r, 8)

    # Discovery Questions
    r = section_header(ws, r, "  DISCOVERY QUESTIONS BY ICP")
    r = spacer(ws, r, 3)
    r = table_header(ws, r, ["", "Segment", "Opening Question", "Follow-up Question", "", "", "", "", "", ""])
    dq = [
        ("", "Enterprise OEM",
         "\"How long does your current simulation setup take from CAD import to first result?\"",
         "\"How many tools does that workflow touch?\""),
        ("", "Consultancy",
         "\"What percentage of your tool licenses are actively used vs. idle at any given time?\"",
         "\"How does licensing cost factor into your project bids?\""),
        ("", "University",
         "\"What simulation tools are your students using, and how accessible are they from home?\"",
         "\"What's your annual software budget for the engineering department?\""),
        ("", "Indie / Startup",
         "\"What simulation tools are you using today, and how are you paying for them?\"",
         "\"How long does it take you to set up and run a basic stress analysis?\""),
    ]
    for i, row_data in enumerate(dq):
        bg = C_STRIPE if i%2 else C_WHITE
        for c, val in enumerate(list(row_data[:4]) + ["","","","","",""], 1):
            bold = c == 2
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=bold, italic=(c > 2), color=C_SLATE if bold else C_BLACK),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left", "center", wrap=True))
        ws.row_dimensions[r].height = 32
        r += 1
    r = spacer(ws, r, 8)

    # Demo Playbook
    r = section_header(ws, r, "03  DEMO PLAYBOOK")
    r = spacer(ws, r, 3)

    r = sub_header(ws, r, "Quick Demo — 15 Minutes (Indie / Inbound)")
    r = table_header(ws, r, ["", "Time", "Stage", "Action", "Goal", "", "", "", "", ""])
    demo_quick = [
        ("", "0–2 min",   "Hook",            "Ask setup time question",
         "Anchor time savings — get their current baseline before showing ODE"),
        ("", "2–8 min",   "Live AI Demo",    "Plain English → AI generates Modelica → solver → results",
         "Show 5-minute simulation live. Demonstrate AI-native, not retrofitted."),
        ("", "8–12 min",  "Platform Tour",   "2 tools max: Geometry + Component",
         "Keep focused on their use case. Don't overwhelm."),
        ("", "12–14 min", "Value Statement", "Quantify delta with their own number",
         "\"You just saw 5 minutes. Your current workflow takes how long?\""),
        ("", "14–15 min", "CTA",             "Free account signup",
         "\"No credit card. First simulation in under 10 minutes.\""),
    ]
    for i, row_data in enumerate(demo_quick):
        bg = C_STRIPE if i%2 else C_WHITE
        for c, val in enumerate(list(row_data[:5]) + ["","","","",""], 1):
            bold = c == 3
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=bold, color=C_INDIGO if c==2 else (C_SLATE if bold else C_BLACK)),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left", "center", wrap=True))
        ws.row_dimensions[r].height = 30
        r += 1
    r = spacer(ws, r, 4)

    r = sub_header(ws, r, "Deep Dive Demo — 45 Minutes (Enterprise)")
    r = table_header(ws, r, ["", "Time", "Stage", "Action", "Goal", "", "", "", "", ""])
    demo_deep = [
        ("", "0–5 min",   "Discovery",        "Specific workflow pain — name their tools",
         "Use their vocabulary. Identify one painful workflow to anchor the demo."),
        ("", "5–10 min",  "Positioning",      "3 differentiators only",
         "AI-native · 12 tools, one platform · Open standards. Max 3 points."),
        ("", "10–30 min", "Custom Demo",      "Pre-prepared model for their industry",
         "Run live. Show AI generation + results interpretation. Domain-relevant."),
        ("", "30–35 min", "Collaboration",    "Real-time co-editing, team comments",
         "Address their collaboration pain. Show PDF-free result sharing."),
        ("", "35–40 min", "ROI Discussion",   "Live ROI calculator with their numbers",
         "Input team size, hours/cycle, license costs. Show 3-year savings."),
        ("", "40–45 min", "Pilot Proposal",   "30-day paid pilot — one real project",
         "\"What project would give you the clearest proof?\""),
    ]
    for i, row_data in enumerate(demo_deep):
        bg = C_STRIPE if i%2 else C_WHITE
        for c, val in enumerate(list(row_data[:5]) + ["","","","",""], 1):
            bold = c == 3
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=bold, color=C_INDIGO if c==2 else (C_SLATE if bold else C_BLACK)),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left", "center", wrap=True))
        ws.row_dimensions[r].height = 30
        r += 1
    r = spacer(ws, r, 8)

    # ROI Framework
    r = section_header(ws, r, "06  ROI FRAMEWORK")
    r = spacer(ws, r, 3)

    r = sub_header(ws, r, "Enterprise ROI — 10-Engineer Team Reference")
    r = table_header(ws, r, ["", "Input / Output", "Value", "Notes", "", "", "", "", "", ""])
    roi = [
        ("", "Engineers using simulation",      "10",           "Reference team size"),
        ("", "Simulation cycles / project",     "10",           "Average complexity"),
        ("", "Hours per cycle (before ODE)",    "8 hrs",        "Benchmark from prospect discovery"),
        ("", "AI setup reduction",              "75%",          "Conservative estimate"),
        ("", "Hourly rate (blended)",           "$100 / hr",    "Engineering team cost"),
        ("", "Projects per year",               "30",           ""),
        ("", "Current tool licenses (annual)",  "$150,000",     "ANSYS + CAD seats"),
        ("", "ODE Team (10 × $199/mo)",         "$23,880",      "Annual ODE cost"),
        ("", "— — —",                          "— — —",        "— — —"),
        ("", "Time saved per cycle",            "6 hours",      "8 hrs × 75%"),
        ("", "Total hours saved / year",        "1,800 hrs",    "6 hrs × 10 cycles × 30 projects"),
        ("", "Time value of savings",           "$180,000",     "1,800 hrs × $100"),
        ("", "License cost reduction",          "$126,120",     "$150,000 − $23,880"),
        ("", "TOTAL ANNUAL VALUE FROM ODE",     "$306,120",     "12.8× ROI vs. ODE cost"),
    ]
    for i, row_data in enumerate(roi):
        bold = "TOTAL" in str(row_data[1]) or row_data[1] == "— — —"
        bg = C_NAVY if "TOTAL" in str(row_data[1]) else (C_RULE if "— — —" in str(row_data[1]) else (C_STRIPE if i%2 else C_WHITE))
        c_txt = C_WHITE if "TOTAL" in str(row_data[1]) else C_BLACK
        for c, val in enumerate(list(row_data[:4]) + ["","","","","",""], 1):
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=bold, color=c_txt),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left", "center", wrap=True))
        ws.row_dimensions[r].height = 20
        r += 1
    r = spacer(ws, r, 4)

    r = sub_header(ws, r, "Competitive TCO — 3-Year Comparison (10-Engineer Team)")
    r = table_header(ws, r, ["", "Tool", "Year 1", "3-Year Total", "AI Included", "", "", "", "", ""])
    tco = [
        ("", "ODE Pro (10 seats)",           "$23,880",   "$71,640",    "Yes — Day 1"),
        ("", "ANSYS Mechanical",              "$300K+",    "$900K+",     "Partial"),
        ("", "Dassault CATIA + SIMULIA",      "$500K+",    "$1.5M+",     "Partial"),
        ("", "SimScale Pro",                  "$48,000",   "$144,000",   "No"),
        ("", "COMSOL Multiphysics",           "$50K–$150K","$150K–450K", "No"),
    ]
    for i, row_data in enumerate(tco):
        ode = row_data[1].startswith("ODE")
        bg = C_INDIGO_LT if ode else (C_STRIPE if i%2 else C_WHITE)
        bold = ode
        for c, val in enumerate(list(row_data[:5]) + ["","","","",""], 1):
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=bold, color=C_INDIGO if (ode and c in (2,3,4)) else C_BLACK),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left" if c==2 else "center", "center"))
        ws.row_dimensions[r].height = 20
        r += 1

    freeze(ws, 4)
    return ws

# ─── SHEET 4: BRAND STRATEGY ─────────────────────────────────────────────────
def build_brand(wb):
    ws = wb.create_sheet("Brand Strategy")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [3, 22, 22, 18, 14, 10, 10, 10, 10, 10])

    r = 2
    r = page_title(ws, r, "Brand Strategy — Positioning, Competitive Landscape & Messaging",
        "ODE · April 2026 · Internal Confidential")
    r = spacer(ws, r, 4)

    kpis = [("12","Engineering Disciplines"),("5","ICP Segments"),
            ("5–50×","Cost Advantage vs. Legacy"),("1","Core Brand Anchor Position")]
    r = kpi_row(ws, r, kpis)
    r = spacer(ws, r, 6)

    # 01 Positioning
    r = section_header(ws, r, "01  BRAND POSITIONING")
    r = spacer(ws, r, 3)

    r = insight_row(ws, r, "Core Position",
        '"Orthogonal is the Dassault of the AI era — the AI-native full-stack engineering platform covering CAD, '
        'MBSE, FEM, CFD, and more, all in the browser, at 5–50× lower cost. From aerospace engineers to Arduino '
        'builders — democratizing serious engineering for everyone."')
    r = spacer(ws, r, 3)

    r = sub_header(ws, r, "Tagline Hierarchy")
    r = table_header(ws, r, ["", "Level", "Tagline", "", "", "", "", "", "", ""])
    taglines = [
        ("", "L1 — Primary Brand",       '"The Dassault of the AI era."'),
        ("", "L2 — Descriptive",         '"AI-native full-stack engineering platform"'),
        ("", "L3 — Value Proposition",   '"Same breadth as Dassault. AI-native from day one. 5–50× lower cost."'),
        ("", "L4 — Competitive Hook",    '"Dassault-grade engineering. Not Dassault-grade pricing."'),
        ("", "L5 — Aspirational",        '"We built the AI-era upgrade Dassault can never ship."'),
        ("", "L6 — Maker Track",         '"LEGO for serious builders. Arduino with AI. 3D print from your browser."'),
    ]
    for i, (a, lvl, tgl) in enumerate(taglines):
        bg = C_STRIPE if i%2 else C_WHITE
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
        for c, val in enumerate(["", lvl, tgl, "","","","","","",""], 1):
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=(c==2), italic=(c==3), color=C_SLATE if c==2 else C_BLACK),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left", "center", wrap=True))
        ws.row_dimensions[r].height = 24
        r += 1
    r = spacer(ws, r, 4)

    r = sub_header(ws, r, "Top 5 Differentiators — Ranked by Impact")
    r = table_header(ws, r, ["", "#", "Differentiator", "Description", "", "", "", "", "", ""])
    diffs = [
        ("", "1", "AI-Native, Not Retrofitted",   "Architecture designed for AI from day one — not bolt-on features"),
        ("", "2", "Unified Engineering Platform", "12 disciplines in one environment — CAD, MBSE, FEM, CFD, RL, Circuit"),
        ("", "3", "Modelica + LLM Synergy",       "Equation-based = natively AI-readable. Structural moat vs. binary formats."),
        ("", "4", "Cloud-Native, Zero-Install",   "Browser-based, real-time collaboration — no desktop install required"),
        ("", "5", "Open Standards First",         "Modelica 4.0, FMI 2.0/3.0, SysML v1.6 — no lock-in, full portability"),
    ]
    for i, row_data in enumerate(diffs):
        bg = C_STRIPE if i%2 else C_WHITE
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
        for c, val in enumerate(list(row_data[:4]) + ["","","","","",""], 1):
            bold = c == 3
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=bold, color=C_INDIGO if c==2 else (C_SLATE if bold else C_BLACK)),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left", "center", wrap=True))
        ws.row_dimensions[r].height = 22
        r += 1
    r = spacer(ws, r, 8)

    # 02 Competitive Landscape
    r = section_header(ws, r, "02  COMPETITIVE LANDSCAPE")
    r = spacer(ws, r, 3)

    r = sub_header(ws, r, "Platform Capability Matrix")
    r = table_header(ws, r, ["", "Platform", "AI-Native", "CAD", "FEM/CFD", "MBSE", "Browser", "Annual Cost / Seat", "Threat", ""])
    comp = [
        ("", "Orthogonal ODE",        "Day 1",      "Yes", "Yes", "Yes", "Yes",   "$948–$1,908",     "—"),
        ("", "Dassault Systèmes",      "Partial",    "Yes", "Yes", "Yes", "No",    "$50K–$500K",      "Medium"),
        ("", "Neural Concept",         "Partial",    "No",  "Yes", "No",  "Yes",   "Undisclosed",     "High"),
        ("", "Autodesk Fusion",        "Retrofitted","Yes", "Yes", "No",  "Partial","$5K–$30K",       "Medium"),
        ("", "SimScale",               "Partial",    "No",  "Yes", "No",  "Yes",   "$2.4K–$4.8K",    "Low"),
        ("", "Onshape",                "Partial",    "Yes", "No",  "No",  "Yes",   "$2.1K–$4.8K",    "Medium"),
    ]
    for i, row_data in enumerate(comp):
        ode = row_data[1].startswith("Orthogonal")
        bg = C_INDIGO_LT if ode else (C_STRIPE if i%2 else C_WHITE)
        for c, val in enumerate(list(row_data[:9]) + [""], 1):
            threat_color = {"High": "C0392B", "Medium": "E67E22", "Low": "27AE60", "—": C_BLACK}.get(val, C_BLACK)
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=ode, color=threat_color if c==9 else (C_INDIGO if ode else C_BLACK)),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("center" if c > 2 else "left", "center"))
        ws.row_dimensions[r].height = 22
        r += 1
    r = spacer(ws, r, 4)

    r = insight_row(ws, r, "Structural Moat",
        "Dassault's binary format is opaque to LLMs. Modelica is equation-based and natively AI-readable. "
        "This is not a feature gap — it is an architecture gap that cannot be closed by bolt-on AI features. "
        "Rebuilding Dassault for AI would destroy backwards compatibility on $5B+ revenue.")
    r = spacer(ws, r, 8)

    # 03 Brand Architecture
    r = section_header(ws, r, "03  BRAND ARCHITECTURE — 12 ENGINEERING TOOLS")
    r = spacer(ws, r, 3)

    r = table_header(ws, r, ["", "Tool", "Phase", "Category", "Description", "", "", "", "", ""])
    tools = [
        ("", "Component", "Phase 1 — Live", "Physics Modelling",    "Modelica AI physics modelling — AI-native"),
        ("", "Geometry",  "Phase 1 — Live", "CAD",                  "Parametric CAD in browser — zero install"),
        ("", "System",    "Phase 1 — Live", "MBSE",                 "MBSE + SysML architecture design"),
        ("", "Vision",    "Phase 1 — Live", "3D Simulation",        "Real-time 3D simulation viewer"),
        ("", "Paper",     "Phase 1 — Live", "Collaboration",        "Collaborative engineering documents"),
        ("", "Interaction","Phase 1 — Live","Parameter Exploration","Live parameter exploration environment"),
        ("", "FEM",       "Phase 2",        "Structural Analysis",  "Finite element analysis"),
        ("", "CFD",       "Phase 2",        "Fluid Dynamics",       "Computational fluid dynamics"),
        ("", "Compute",   "Phase 3",        "Scientific Computing", "High-performance scientific computing"),
        ("", "Optim",     "Phase 3",        "Optimisation",         "Multi-objective design optimisation"),
        ("", "RL",        "Phase 4",        "Machine Learning",     "Reinforcement learning environment"),
        ("", "Circuit",   "Phase 4",        "Electronics",          "KiCad PCB + hardware/software co-simulation"),
    ]
    for i, row_data in enumerate(tools):
        live = "Live" in row_data[2]
        bg = C_INDIGO_LT if live else (C_STRIPE if i%2 else C_WHITE)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=9)
        for c, val in enumerate(list(row_data[:5]) + ["","","","",""], 1):
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=(c==2), color=C_INDIGO if (live and c==3) else (C_SLATE if c==2 else C_BLACK)),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left", "center", wrap=True))
        ws.row_dimensions[r].height = 20
        r += 1
    r = spacer(ws, r, 8)

    # 04 Messaging Matrix
    r = section_header(ws, r, "04  MESSAGING MATRIX BY ICP")
    r = spacer(ws, r, 3)

    r = table_header(ws, r, ["", "Segment", "Core Message", "Key Hook", "Proof Point", "", "", "", "", ""])
    msg = [
        ("", "Enterprise OEMs",
         '"Replace your Dassault stack at 10% of the cost"',
         "$50K–$500K/seat with zero AI-era value; multi-year lock-in",
         "AI-native FEM + CFD in token plan; full Dassault suite equivalent"),
        ("", "Engineering Consultancies",
         '"Token-based: pay per simulation run, not per seat-year"',
         "License utilization 30–50% idle; tool costs 15–25% of project budget",
         "ODE FEM + CFD replaces ANSYS + SolidWorks in one subscription"),
        ("", "Universities",
         '"Free for students, affordable for departments"',
         "80%+ academic discount; browser-based = zero IT overhead",
         "Graduate job-ready engineers; research-grade simulation in browser"),
        ("", "Indie Engineers",
         '"Enterprise-grade simulation for $49/month"',
         "ANSYS single seat costs more than a seed-stage monthly cloud budget",
         "AI does the meshing; first simulation in 60 seconds"),
        ("", "Maker / Enthusiast",
         '"LEGO for serious builders. Arduino with AI. 3D print from your browser."',
         "Fragmented tools: Tinkercad + Arduino IDE + Cura + Shapeways — no single platform",
         "Maker CAD + Firmware IDE + 3D print in one $9–$29/mo flow"),
    ]
    for i, row_data in enumerate(msg):
        bg = C_STRIPE if i%2 else C_WHITE
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=9)
        for c, val in enumerate(list(row_data[:5]) + ["","","","",""], 1):
            bold = c == 2
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=bold, italic=(c==3),
                    color=C_SLATE if bold else C_BLACK),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left", "center", wrap=True))
        ws.row_dimensions[r].height = 42
        r += 1

    freeze(ws, 4)
    return ws

# ─── SHEET 5: MARKETING STRATEGY ────────────────────────────────────────────
def build_marketing(wb):
    ws = wb.create_sheet("Marketing Strategy")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [3, 20, 16, 18, 16, 12, 12, 12, 12, 10])

    r = 2
    r = page_title(ws, r, "Marketing Strategy — Content, Channels, SEO & Webinar Calendar",
        "ODE · April 2026 · Internal Confidential")
    r = spacer(ws, r, 4)

    kpis = [("5+","Content Pieces / Week"),("25K","Organic Traffic / mo (Q4)"),
            ("8,000","Email List Target (Q4)"),("120","MQLs / Month (Q4)"),("$7.5K","Monthly Paid Budget")]
    r = kpi_row(ws, r, kpis)
    r = spacer(ws, r, 6)

    # 01 Content Engine
    r = section_header(ws, r, "01  CONTENT ENGINE")
    r = spacer(ws, r, 3)

    r = insight_row(ws, r, "Execution Rule",
        "No human writes a cold first draft. Claude generates; AI Marketing (0.5 FTE) refines in under 15 minutes. "
        "5 content pieces per week, every week, from Day 1. AI output tracked as a team KPI.")
    r = spacer(ws, r, 3)

    r = sub_header(ws, r, "Weekly Content Cadence by Channel")
    r = table_header(ws, r, ["", "Channel", "Frequency", "AI Tool", "Human Review", "Goal", "", "", "", ""])
    content = [
        ("", "LinkedIn (Company)",  "5× / week",    "Claude",          "<10 min/post",
         "Mon: thought leadership · Tue: demo · Wed: story · Thu: technical · Fri: culture"),
        ("", "LinkedIn (CEO)",      "3× / week",    "Claude (voice-matched)", "Voice-match + approve",
         "Founder narrative — \"Dassault of the AI era\""),
        ("", "Technical Blog",      "2× / week",    "Claude + Perplexity", "<30 min/post",
         "FEM, CFD, MBSE comparisons vs. legacy tools"),
        ("", "Email Newsletter",    "1× / week",    "Claude + Customer.io", "Approve and send",
         "ICP-segmented nurture sequence"),
        ("", "YouTube",             "2× / month",   "Runway + Descript",   "Light edit",
         "Product demos, simulation walkthroughs"),
        ("", "SEO Landing Pages",   "2× / month",   "Cursor + Framer",     "QA and publish",
         "Competitor comparison pages (\"ODE vs. ANSYS\")"),
        ("", "Twitter / X",         "3–5× / week",  "Claude",              "Quick review",
         "Community engagement, Modelica/AI ecosystem"),
    ]
    for i, row_data in enumerate(content):
        bg = C_STRIPE if i%2 else C_WHITE
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=9)
        for c, val in enumerate(list(row_data[:6]) + ["","","",""], 1):
            bold = c == 2
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=bold, color=C_SLATE if bold else C_BLACK),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left", "center", wrap=True))
        ws.row_dimensions[r].height = 28
        r += 1
    r = spacer(ws, r, 8)

    # 02 Channel Mix
    r = section_header(ws, r, "02  PAID CHANNEL MIX — $7,500 / MONTH")
    r = spacer(ws, r, 3)

    r = table_header(ws, r, ["", "Channel", "Budget / mo", "% of Total", "Primary Target", "KPI", "", "", "", ""])
    channels = [
        ("", "Google Search Ads",   "$3,000", "40%",   "High-intent: \"AI simulation\", \"AI CAD\"",    "CPC <$3 · Conv rate >3%"),
        ("", "LinkedIn Sponsored",  "$2,000", "26.7%", "Engineering Mgrs, VP Eng, CTO (500+ employees)","CPL $15–$25"),
        ("", "Sponsorships",        "$1,500", "20%",   "Engineering podcasts, Modelica conferences",     "Brand impression"),
        ("", "Retargeting",         "$1,000", "13.3%", "Website visitors & trial users",                 "CPA <$200"),
        ("", "Total",               "$7,500", "100%",  "—",                                              "Blended CAC <$800"),
    ]
    for i, row_data in enumerate(channels):
        bold = row_data[1] == "Total"
        bg = C_NAVY if bold else (C_STRIPE if i%2 else C_WHITE)
        c_txt = C_WHITE if bold else C_BLACK
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
        for c, val in enumerate(list(row_data[:6]) + ["","","",""], 1):
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=bold, color=c_txt),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left" if c in (2,5,6) else "center", "center"))
        ws.row_dimensions[r].height = 22
        r += 1
    r = spacer(ws, r, 8)

    # 04 Email Sequences
    r = section_header(ws, r, "04  EMAIL SEQUENCES — MAKER ONBOARDING (DAY 0–21)")
    r = spacer(ws, r, 3)

    r = table_header(ws, r, ["", "Day", "Email Title", "Goal", "Key Content", "Conversion Milestone", "", "", "", ""])
    emails = [
        ("", "Day 0",  "Welcome to ODE: Your All-in-One Maker Platform",
         "Activate sign-up", "Explain Maker CAD + Firmware IDE + Print", "Account activated"),
        ("", "Day 1",  "Open Maker CAD + start a template project",
         "First onboarding", "3 starter templates: LEGO bracket, RPi case, Arduino enclosure", "First project open"),
        ("", "Day 3",  "Your first LEGO CAD design — 10 minutes, browser only",
         "First product use","Complete first parametric LEGO design — GIF walkthrough + video link", "Design created"),
        ("", "Day 7",  "Your AI firmware co-pilot is ready — flash your first Arduino",
         "Firmware IDE intro","Describe blink pattern in plain English → AI writes firmware → one-click flash", "Firmware IDE opened"),
        ("", "Day 12", "Turn your design into reality — $5 print credit inside",
         "Monetization intro","One-click STL → print order; delivery time + material options", "Print order placed"),
        ("", "Day 21", "Unlock unlimited builds — Maker Pro for $29/mo",
         "Conversion CTA",   "21-day usage summary; compare vs. Tinkercad + Arduino Cloud costs", "Upgrade → paid"),
    ]
    for i, row_data in enumerate(emails):
        bg = C_STRIPE if i%2 else C_WHITE
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
        for c, val in enumerate(list(row_data[:6]) + ["","","",""], 1):
            bold = c == 3
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=bold, color=C_INDIGO if c==2 else (C_SLATE if bold else C_BLACK)),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("left", "center", wrap=True))
        ws.row_dimensions[r].height = 32
        r += 1
    r = spacer(ws, r, 8)

    # 05 SEO
    r = section_header(ws, r, "05  SEO STRATEGY — TARGET: 25,000 ORGANIC VISITS / MO (Q4)")
    r = spacer(ws, r, 3)

    r = table_header(ws, r, ["", "Keyword", "Monthly Volume", "Difficulty", "Intent", "", "", "", "", ""])
    seo = [
        ("", "AI CAD software",                  "2,400", "Medium",   "High purchase"),
        ("", "AI simulation software",            "1,800", "Medium",   "High purchase"),
        ("", "AI engineering platform",           "1,300", "Low–Med",  "High purchase"),
        ("", "Browser-based CAD",                 "1,100", "Medium",   "High purchase"),
        ("", "Cloud engineering tools",           "2,000", "High",     "Medium"),
        ("", "AI-powered simulation",             "600",   "Low",      "High purchase"),
        ("", "Modelica simulation online",        "500",   "Low",      "High purchase"),
        ("", "Cloud engineering simulation tool", "260",   "Low",      "High purchase"),
    ]
    for i, row_data in enumerate(seo):
        bg = C_STRIPE if i%2 else C_WHITE
        for c, val in enumerate(list(row_data[:5]) + ["","","","",""], 1):
            diff_color = {"High": "C0392B", "Medium": "E67E22", "Low": "27AE60",
                         "Low–Med": "F39C12"}.get(val, C_BLACK)
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, color=diff_color if c==4 else C_BLACK),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("center" if c in (3,4,5) else "left", "center"))
        ws.row_dimensions[r].height = 20
        r += 1
    r = spacer(ws, r, 8)

    # 06 Webinar Calendar
    r = section_header(ws, r, "06  WEBINAR CALENDAR — Q2–Q3 2026")
    r = spacer(ws, r, 3)

    r = table_header(ws, r, ["", "Date", "Title", "Target ICP", "Series", "", "", "", "", ""])
    webinars = [
        ("", "Apr 15", "ODE 101: Your First AI-Assisted Simulation in 15 Minutes",     "Indie Engineers · Universities",     "Enterprise Series"),
        ("", "Apr 29", "Enterprise Simulation at Scale: Prototype to Production",        "Enterprise OEMs",                    "Enterprise Series"),
        ("", "May 13", "Modelica + Claude: Building Multi-Domain Models with AI",        "Consultancies · Universities",       "Enterprise Series"),
        ("", "May 27", "Migrating from MATLAB/Simulink to ODE — A 30-Day Guide",        "Enterprise OEMs · Consultancies",    "Enterprise Series"),
        ("", "Jun 10", "Teaching Systems Engineering with AI: A Professor's Toolkit",    "Universities",                       "Enterprise Series"),
        ("", "Jun 24", "ODE for Automotive: Thermal, Electrical & Powertrain Co-sim",   "Enterprise OEMs",                    "Enterprise Series"),
        ("", "Jul 8",  "The Solo Engineer's Toolkit: ODE Replaces 5 Desktop Tools",     "Indie Engineers",                    "Enterprise Series"),
        ("", "Jul 15", "Build your first AI-designed 3D print in 30 minutes",           "LEGO hobbyists · 3D printing",       "Maker Series"),
        ("", "Jul 22", "Building Digital Twins with ODE and FMI/FMU",                   "Enterprise OEMs · Consultancies",    "Enterprise Series"),
        ("", "Aug 5",  "AI-Assisted Design Optimisation: Concept to Converged",         "All ICPs",                           "Enterprise Series"),
        ("", "Aug 12", "Arduino firmware in 10 minutes — AI writes the code",           "Arduino devs · RPi makers",          "Maker Series"),
        ("", "Sep 2",  "Customer Showcase: Cut Simulation Setup Time by 90%",           "Enterprise OEMs",                    "Enterprise Series"),
        ("", "Sep 9",  "LEGO modular enclosures: CAD + 3D print + firmware in one",     "LEGO hobbyists · Makers",            "Maker Series"),
        ("", "Sep 16", "Open Standards in Practice: Interoperability and Data Freedom",  "All ICPs",                           "Enterprise Series"),
    ]
    for i, row_data in enumerate(webinars):
        maker = "Maker Series" in row_data[4]
        bg = C_INDIGO_LT if maker else (C_STRIPE if i%2 else C_WHITE)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        for c, val in enumerate(list(row_data[:5]) + ["","","","",""], 1):
            bold = c == 2
            apply(ws, r, c, val,
                fnt=font("Calibri", 9, bold=bold,
                    color=C_INDIGO if (maker and c==5) else (C_SLATE if bold else C_BLACK)),
                fll=fill(bg), brd=border("thin", C_RULE),
                aln=align("center" if c==2 else "left", "center", wrap=True))
        ws.row_dimensions[r].height = 24
        r += 1

    freeze(ws, 4)
    return ws

# ─── MAIN ────────────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_summary(wb)
    build_gtm(wb)
    build_sales(wb)
    build_brand(wb)
    build_marketing(wb)

    out = "/Users/dennis/Project/ODE_G2M/deck/ODE_GTM_Strategy_2026.xlsx"
    wb.save(out)
    print(f"✓  Saved: {out}")

if __name__ == "__main__":
    main()
